import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import calendar
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkcalendar import Calendar
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pulp import LpMaximize, LpProblem, LpVariable, lpSum, LpStatus, LpMinimize, value


class Employee:
    def __init__(self, name, seniority, emp_id):
        self.name = name
        self.seniority = seniority
        self.id = emp_id
        self.schedule = []
        self.training_days = []

    def add_shift(self, date, shift):
        self.schedule.append((date, shift))

    def get_shifts(self):
        return self.schedule

    def add_training_day(self, date):
        self.training_days.append(date)


class ShiftPreference:
    def __init__(self, emp_id, preferences):
        self.emp_id = emp_id
        self.preferences = preferences
        # Dictionary of date to preferred shift


class Schedule:
    def __init__(self):
        self.employees = []
        self.schedules = {}
        self.shift_preferences = []

    def load_employees_from_json(self, file_path):
        with open(file_path, 'r') as file:
            employee_data = json.load(file)
            self.employees = [Employee(emp['name'], emp['seniority'], emp['id']) for emp in employee_data]

    def collect_shift_preferences(self, year, month, preferences):
        self.shift_preferences = preferences

    def generate_schedule(self, year, month, gazetted_holidays):
        shifts = ['M', 'E', 'N', 'Off', 'TR']
        days_in_month = calendar.monthrange(year, month)[1]

        # Calculate weekends and total holidays
        weekends = sum(1 for d in range(1, days_in_month + 1) if calendar.weekday(year, month, d) >= 5)
        total_holidays = weekends + gazetted_holidays

        print(f"Total holidays: {total_holidays}, total weekends: {weekends}")

        self.schedules[(year, month)] = pd.DataFrame(index=[e.name for e in self.employees],
                                                     columns=range(1, days_in_month + 1))

        # Create the LP problem
        prob = LpProblem("ShiftScheduling", LpMinimize)

        # Variables
        shifts_vars = LpVariable.dicts("Shift",
                                       ((e.name, d, s) for e in self.employees for d in range(1, days_in_month + 1) for
                                        s in shifts),
                                       cat='Binary')

        # Objective: Minimize deviations from preferences (assuming preferences are given in the form of penalties)
        prob += lpSum(
            shifts_vars[e.name, d, s] for e in self.employees for d in range(1, days_in_month + 1) for s in shifts if
            self.get_preference_penalty(e.id, datetime(year, month, d), s) > 0)

        # Constraints
        for d in range(1, days_in_month + 1):
            # At least 1 employees in morning (M) and evening (E) shifts, and exactly 2 in night (N) shifts
            prob += lpSum(shifts_vars[e.name, d, 'M'] for e in self.employees) >= 1
            prob += lpSum(shifts_vars[e.name, d, 'E'] for e in self.employees) >= 1
            prob += lpSum(shifts_vars[e.name, d, 'N'] for e in self.employees) == 2

            # Each employee can have only one shift per day
            for e in self.employees:
                prob += lpSum(shifts_vars[e.name, d, s] for s in shifts) == 1
                # after a night shift, can have a leave or another night shift followed by two leaves
                if d == 1:
                    continue
                for s in ['M', 'E']:
                    prob += lpSum([shifts_vars[e.name, d, s], shifts_vars[e.name, d - 1, 'N']]) <= 1
                # cant have a morning shift (or TR) after an evening shift
                prob += lpSum([shifts_vars[e.name, d, 'M'], shifts_vars[e.name, d - 1, 'E']]) <= 1
                prob += lpSum([shifts_vars[e.name, d, 'TR'], shifts_vars[e.name, d - 1, 'E']]) <= 1
                if d <= 2:
                    continue
                for s in ['M', 'E', 'N', 'TR']:
                    prob += lpSum([shifts_vars[e.name, d, s], shifts_vars[e.name, d - 1, 'N'],
                                   shifts_vars[e.name, d - 2, 'N']]) <= 2
                    prob += lpSum([shifts_vars[e.name, d, s], shifts_vars[e.name, d - 1, 'TR'],
                                   shifts_vars[e.name, d - 2, 'N']]) <= 2
                if d <= 3:
                    continue
                for s in ['M', 'E', 'N', 'TR']:
                    prob += lpSum([shifts_vars[e.name, d, s], shifts_vars[e.name, d - 2, 'N'],
                                   shifts_vars[e.name, d - 3, 'N']]) <= 2
                    prob += lpSum([shifts_vars[e.name, d, s], shifts_vars[e.name, d - 2, 'TR'],
                                   shifts_vars[e.name, d - 3, 'N']]) <= 2

                # ensure that each employee gets a maximum of 6 consecutive shifts
                # i.e. they have a holiday at-least once in 7 days
                if d <= 6:
                    continue
                prob += lpSum([shifts_vars[e.name, d, 'Off'], shifts_vars[e.name, d - 1, 'Off'],
                               shifts_vars[e.name, d - 2, 'Off'], shifts_vars[e.name, d - 3, 'Off'],
                               shifts_vars[e.name, d - 4, 'Off'], shifts_vars[e.name, d - 5, 'Off'],
                               shifts_vars[e.name, d - 6, 'Off']]) >= 1

        # Ensure that each employee gets equal number of night shifts (+-1)
        avg_night_shifts = 2 * days_in_month / len(self.employees)
        for e in self.employees:
            prob += lpSum(shifts_vars[e.name, d, 'N'] for d in range(1, days_in_month + 1)) >= avg_night_shifts - 1
            prob += lpSum(shifts_vars[e.name, d, 'N'] for d in range(1, days_in_month + 1)) <= avg_night_shifts + 1

        # Ensure that each employee gets equal total number of shifts (+-1)
        avg_total_shifts = days_in_month - total_holidays
        for e in self.employees:
            prob += lpSum(shifts_vars[e.name, d, s] for d in range(1, days_in_month + 1) for s in shifts if
                          s != 'Off') >= avg_total_shifts - 1
            prob += lpSum(shifts_vars[e.name, d, s] for d in range(1, days_in_month + 1) for s in shifts if
                          s != 'Off') <= avg_total_shifts + 1

        # Ensure that each employee gets the correct number of holidays
        # TODO: RELAX CONSTRAINTS IF PROBLEM BECOMES INFEASIBLE
        for e in self.employees:
            prob += lpSum(shifts_vars[e.name, d, 'Off'] for d in range(1, days_in_month + 1)) >= total_holidays - 1
            prob += lpSum(shifts_vars[e.name, d, 'Off'] for d in range(1, days_in_month + 1)) <= total_holidays + 1

        # Solve the problem
        prob.solve()

        # Check if the problem is infeasible
        if LpStatus[prob.status] == 'Infeasible':
            # Special case: handle infeasibility if needed

            return ()

        # Apply the solution
        for d in range(1, days_in_month + 1):
            for e in self.employees:
                for s in shifts:
                    if value(shifts_vars[e.name, d, s]) == 1:
                        self.schedules[(year, month)].at[e.name, d] = s
                        e.add_shift(datetime(year, month, d), s)

        print(f"Schedule generated successfully for {calendar.month_name[month]} {year}.")

    def get_preference_penalty(self, emp_id, date, shift):
        for preference in self.shift_preferences:
            if preference.emp_id == emp_id:
                if date.strftime('%Y-%m-%d') in preference.preferences:
                    if preference.preferences[date.strftime('%Y-%m-%d')] == shift:
                        return 0
                    else:
                        return 1
        return 1

    def display_schedule(self, year, month):
        if (year, month) in self.schedules:
            print(f""add Company Name here"")
            print(f"Duty Roster for {calendar.month_name[month]} {year}")
            print(self.schedules[(year, month)].T.fillna('Off'))
        else:
            print(f"No schedule found for {calendar.month_name[month]} {year}")

    def employee_performance(self):
        performance = {}
        for employee in self.employees:
            performance[employee.name] = {
                'M': len([s for s in employee.get_shifts() if s[1] == 'M']),
                'E': len([s for s in employee.get_shifts() if s[1] == 'E']),
                'N': len([s for s in employee.get_shifts() if s[1] == 'N']),
                'Off': len([s for s in employee.get_shifts() if s[1] == 'Off']),
                'TR': len([s for s in employee.get_shifts() if s[1] == 'TR']),
            }
        return performance


class ScheduleApp(tk.Tk):
    def __init__(self, schedule):
        super().__init__()
        self.gazetted_holidays = tk.IntVar()  # Define as IntVar
        self.schedule = schedule
        self.title("Duty Roster Scheduler")
        self.geometry("800x600")

        self.employee_file_path = tk.StringVar()
        self.selected_year = tk.StringVar()  # No default value
        self.selected_month = tk.StringVar()  # No default value

        self.create_widgets()

    def create_widgets(self):
        frame = ttk.Frame(self)
        frame.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(frame, text="Load Employees JSON File:").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Entry(frame, textvariable=self.employee_file_path, width=50).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="Browse", command=self.browse_file).pack(side=tk.LEFT)
        ttk.Button(frame, text="Load", command=self.load_employees).pack(side=tk.LEFT)

        self.employee_table = ttk.Treeview(self, columns=("Name", "Seniority", "ID"), show='headings')
        self.employee_table.heading("Name", text="Name")
        self.employee_table.heading("Seniority", text="Seniority")
        self.employee_table.heading("ID", text="ID")
        self.employee_table.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.shift_preferences = []
        self.create_preference_widgets()

        # Year and Month selection
        date_frame = ttk.Frame(self)
        date_frame.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(date_frame, text="Select Year:").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(date_frame, textvariable=self.selected_year, width=10).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(date_frame, text="Select Month:").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(date_frame, textvariable=self.selected_month, width=10).pack(side=tk.LEFT, padx=(0, 10))

        # Gazetted Holiday
        ttk.Label(date_frame,
                  text="Select No. of Gazetted Holidays in that Month (That does not lie on a weekend) :").pack(
            side=tk.LEFT,
            padx=(0, 8))
        ttk.Entry(date_frame, textvariable=self.gazetted_holidays, width=10).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Button(self, text="Generate Schedule", command=self.generate_schedule).pack(pady=(10, 0))

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        self.employee_file_path.set(file_path)

    def load_employees(self):
        file_path = self.employee_file_path.get()
        self.schedule.load_employees_from_json(file_path)
        self.update_employee_table()

    def update_employee_table(self):
        for row in self.employee_table.get_children():
            self.employee_table.delete(row)

        for employee in self.schedule.employees:
            self.employee_table.insert("", "end", values=(employee.name, employee.seniority, employee.id))

    def create_preference_widgets(self):
        frame = ttk.Frame(self)
        frame.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(frame, text="Shift Preferences:").pack(side=tk.LEFT, padx=(0, 10))

        ttk.Button(frame, text="Add Preference", command=self.add_preference).pack(side=tk.LEFT)

    def add_preference(self):
        preference_window = tk.Toplevel(self)
        preference_window.title("Add Shift Preference")
        preference_window.geometry("400x300")

        cal_frame = ttk.Frame(preference_window)
        cal_frame.pack(padx=10, pady=10)

        ttk.Label(cal_frame, text="Select Date:").pack()
        cal = Calendar(cal_frame, selectmode="day", date_pattern="yyyy-mm-dd")
        cal.pack()

        ttk.Label(preference_window, text="Select Employee:").pack(padx=10, pady=(10, 0))
        emp_var = tk.StringVar()
        emp_dropdown = ttk.Combobox(preference_window, textvariable=emp_var, state="readonly")
        emp_dropdown.pack(padx=10, pady=(0, 10))
        emp_dropdown['values'] = [employee.name for employee in self.schedule.employees]

        ttk.Label(preference_window, text="Select Shift:").pack(padx=10, pady=(10, 0))
        shift_var = tk.StringVar()
        shift_dropdown = ttk.Combobox(preference_window, textvariable=shift_var, state="readonly")
        shift_dropdown.pack(padx=10, pady=(0, 10))
        shift_dropdown['values'] = ['M', 'E', 'N', 'Off', 'TR']

        ttk.Button(preference_window, text="Save Preference",
                   command=lambda: self.save_preference(cal, emp_var.get(), shift_var.get(), preference_window)).pack()

    def save_preference(self, cal, employee_name, shift, preference_window):
        selected_date = cal.get_date()
        emp_id = next((emp.id for emp in self.schedule.employees if emp.name == employee_name), None)
        if emp_id:
            preference = ShiftPreference(emp_id, {selected_date: shift})
            self.shift_preferences.append(preference)

    def generate_schedule(self):
        if not self.schedule.employees:
            messagebox.showwarning("Warning", "Please load employees first.")
            return

        try:
            year = int(self.selected_year.get())
            month = int(self.selected_month.get())
            gazetted_holidays = int(self.gazetted_holidays.get())  # Retrieve value from IntVar
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Warning", "Please enter a valid year and month.")
            return

        print(f"Year: {year}, Month: {month}, Gazetted Holidays: {gazetted_holidays}")  # Debug print

        self.schedule.collect_shift_preferences(year, month, self.shift_preferences)
        self.schedule.generate_schedule(year, month, gazetted_holidays)
        self.schedule.display_schedule(year, month)
        self.show_schedule(year, month)

    def show_schedule(self, year, month):
        if (year, month) in self.schedule.schedules:
            schedule_window = tk.Toplevel(self)
            schedule_window.title(f"Duty Roster for {calendar.month_name[month]} {year}")
            schedule_window.geometry("900x700")

            schedule_frame = ttk.Frame(schedule_window)
            schedule_frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=False)

            schedule_table = ttk.Treeview(schedule_frame, columns=[str(day) for day in
                                                                   range(1, calendar.monthrange(year, month)[1] + 1)],
                                          show='headings')
            for day in range(1, calendar.monthrange(year, month)[1] + 1):
                schedule_table.heading(str(day), text=str(day))

            for employee in self.schedule.employees:
                schedule_table.insert("", "end", values=[employee.name] + [
                    self.schedule.schedules[(year, month)].at[employee.name, day] if not pd.isnull(
                        self.schedule.schedules[(year, month)].at[employee.name, day]) else 'Off' for day in
                    range(1, calendar.monthrange(year, month)[1] + 1)])

            schedule_table.pack(padx=20, pady=20, fill=tk.BOTH, expand=False)

            ttk.Button(schedule_window, text="Export Schedule",
                       command=lambda: self.export_schedule(year, month)).pack()

        else:
            messagebox.showwarning("Warning", f"No schedule found for {calendar.month_name[month]} {year}.")

    def export_schedule(self, year, month):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        colors = {
            'Off': 'FF0000',  # Red
            'M': 'FFCCCB',  # Coral
            'E': 'CCFFFF',  # Cyan
            'TR': 'FFCC99',  # Orange
            'N': 'CCFFCC'  # Mint
        }

        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Duty Roster {calendar.month_name[month]} {year}"

            # Header
            ws.append(['Employee'] + [str(day) for day in range(1, calendar.monthrange(year, month)[1] + 1)])

            # Data
            for employee in self.schedule.employees:
                ws.append([employee.name] + [
                    self.schedule.schedules[(year, month)].at[employee.name, day] if not pd.isnull(
                        self.schedule.schedules[(year, month)].at[employee.name, day]) else 'Off' for day in
                    range(1, calendar.monthrange(year, month)[1] + 1)])

            # Apply colors to cells based on the shift type
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                for cell in row:
                    shift_type = cell.value
                    if shift_type in colors:
                        cell.fill = PatternFill(start_color=colors[shift_type], end_color=colors[shift_type],
                                                fill_type="solid")

            wb.save(file_path)
            messagebox.showinfo("Info", "Schedule exported successfully!")


if __name__ == "__main__":
    schedule = Schedule()
    app = ScheduleApp(schedule)
    app.mainloop()


('\n'
 'Check last 2 days of each employee and according to it assign the Shift for next month.\n'
 '\n'
 'Also sometimes the shifts assigned maybe more than 5-6 days,\n'
 'In that case check for Tr in that day and reassign shifts accordingly.\n'
 '\n'
 'Also Shift preference are not constraints, used as options, can be discarded if schedule unable to form\n'
 '\n'
 'Holidays in a month is +-1 (as of now, change here if changed above), so Tr can also act as a Holiday for that Employee\n'
 '(If needed)\n'
 '\n'
 'Tr if not training, can also be used as an option for employee to pick from Morning or Evening Shift\n'
 'Evening only when it comes before "E"\n'
 'And can also be used in case of Paid Leaves\n')

 #Made by Ankit Das and Khwiash Mankotia
